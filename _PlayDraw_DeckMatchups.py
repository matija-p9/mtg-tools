import openpyxl
excel_file = openpyxl.load_workbook("./reference/MTG Pula - Liga 2024 #1.xlsx")

deck_name = "Amulet Titan"
date_list = ["2024-05-22", "2024-05-15", "2024-05-08", "2024-04-24"]

# Prepare empty dictionary to populate.
matchup_matrix = {}
for date in date_list:
    # Assign the correctly dated sheet.
    ws = excel_file[date]
    i = 1   # Iterator for rows reference.
    for row in ws:
        deck_left = str(ws['B'+str(i)].value)
        deck_right = str(ws['E'+str(i)].value)
        # Remove the brackets for minor deck differences.
        if deck_left.find('(') != -1:
            deck_left = deck_left[:deck_left.find('(')-1]
        if deck_right.find('(') != -1:
            deck_right = deck_right[:deck_right.find('(')-1]

        # Check only if matchup IS NOT mirror.
        if deck_left != deck_right:
            
            if deck_left == deck_name:
                # Count deck left-hand side wins.
                if ws['C'+str(i)].value > ws['D'+str(i)].value and ws['B'+str(i)].font.b:
                    try: # If entry exists, add to OTP, then add to total.
                        matchup_matrix[deck_right+" (OTP)"][0] += 1
                        matchup_matrix[deck_right][0] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_right+" (OTP)"] = [1, 0]
                        try: matchup_matrix[deck_right][0] += 1
                        except: matchup_matrix[deck_right] = [1, 0]
                elif ws['C'+str(i)].value > ws['D'+str(i)].value and ws['E'+str(i)].font.b:
                    try: # If entry exists, add to OTD, then add to total.
                        matchup_matrix[deck_right+" (OTD)"][0] += 1
                        matchup_matrix[deck_right][0] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_right+" (OTD)"] = [1, 0]
                        try: matchup_matrix[deck_right][0] += 1
                        except: matchup_matrix[deck_right] = [1, 0]
                elif ws['C'+str(i)].value > ws['D'+str(i)].value:
                    # If no P/D data, just add to total.
                    try: matchup_matrix[deck_right][0] += 1
                    except: matchup_matrix[deck_right] = [1, 0]

                # Count deck left-hand side losses.
                if ws['C'+str(i)].value < ws['D'+str(i)].value and ws['B'+str(i)].font.b:
                    try: # If entry exists, add to OTP, then add to total.
                        matchup_matrix[deck_right+" (OTP)"][1] += 1
                        matchup_matrix[deck_right][1] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_right+" (OTP)"] = [0, 1]
                        try: matchup_matrix[deck_right][1] += 1
                        except: matchup_matrix[deck_right] = [0, 1]
                elif ws['C'+str(i)].value < ws['D'+str(i)].value and ws['E'+str(i)].font.b:
                    try: # If entry exists, add to OTD, then add to total.
                        matchup_matrix[deck_right+" (OTD)"][1] += 1
                        matchup_matrix[deck_right][1] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_right+" (OTD)"] = [0, 1]
                        try: matchup_matrix[deck_right][1] += 1
                        except: matchup_matrix[deck_right] = [0, 1]
                elif ws['C'+str(i)].value < ws['D'+str(i)].value:
                    # If no P/D data, just add to total.
                    try: matchup_matrix[deck_right][1] += 1
                    except: matchup_matrix[deck_right] = [0, 1]

            if deck_right == deck_name:
                # Count deck right-hand side wins.
                if ws['C'+str(i)].value < ws['D'+str(i)].value and ws['B'+str(i)].font.b:
                    try: # If entry exists, add to OTD, then add to total.
                        matchup_matrix[deck_right+" (OTD)"][0] += 1
                        matchup_matrix[deck_right][0] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_right+" (OTD)"] = [1, 0]
                        try: matchup_matrix[deck_right][0] += 1
                        except: matchup_matrix[deck_right] = [1, 0]
                elif ws['C'+str(i)].value < ws['D'+str(i)].value and ws['E'+str(i)].font.b:
                    try: # If entry exists, add to OTP, then add to total.
                        matchup_matrix[deck_right+" (OTP)"][0] += 1
                        matchup_matrix[deck_right][0] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_right+" (OTP)"] = [1, 0]
                        try: matchup_matrix[deck_right][0] += 1
                        except: matchup_matrix[deck_right] = [1, 0]
                elif ws['C'+str(i)].value < ws['D'+str(i)].value:
                    # If no P/D data, just add to total.
                    try: matchup_matrix[deck_right][0] += 1
                    except: matchup_matrix[deck_right] = [1, 0]

                # Count deck right-hand side losses.
                if ws['C'+str(i)].value > ws['D'+str(i)].value and ws['B'+str(i)].font.b:
                    try: # If entry exists, add to OTD, then add to total.
                        matchup_matrix[deck_left+" (OTD)"][1] += 1
                        matchup_matrix[deck_left][1] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_left+" (OTD)"] = [0, 1]
                        try: matchup_matrix[deck_left][1] += 1
                        except: matchup_matrix[deck_left] = [0, 1]
                elif ws['C'+str(i)].value > ws['D'+str(i)].value and ws['E'+str(i)].font.b:
                    try: # If entry exists, add to OTP, then add to total.
                        matchup_matrix[deck_left+" (OTP)"][1] += 1
                        matchup_matrix[deck_left][1] += 1
                    except: # If not, create it, try adding to total, if not, create total.
                        matchup_matrix[deck_left+" (OTP)"] = [0, 1]
                        try: matchup_matrix[deck_left][1] += 1
                        except: matchup_matrix[deck_left] = [0, 1]
                elif ws['C'+str(i)].value > ws['D'+str(i)].value:
                        # If no P/D data, just add to total.
                        try: matchup_matrix[deck_left][1] += 1
                        except: matchup_matrix[deck_left] = [0, 1]
        i += 1

# Sort data alphabetically.
matchup_matrix = dict(sorted(matchup_matrix.items()))

for opponent in matchup_matrix:
    print(opponent, matchup_matrix[opponent])